import os
import numpy as np
from openpyxl import Workbook
from config import *

def show_mask(mask, ax, random_color=False):
    if random_color:
        color = np.concatenate([np.random.random(3), np.array([0.6])], axis=0)
    else:
        color = np.array([30 / 255, 144 / 255, 255 / 255, 0.6])
    h, w = mask.shape[-2:]
    mask_image = mask.reshape(h, w, 1) * color.reshape(1, 1, -1)
    ax.imshow(mask_image)


def show_points(coords, labels, ax, marker_size=50):
    pos_points = coords[labels == 1]
    neg_points = coords[labels == 0]

    ax.scatter(pos_points[:, 0], pos_points[:, 1], color=GREEN_COLOR, marker='*', s=marker_size, edgecolor='white',
            linewidth=LINEWIDTH)
    ax.scatter(neg_points[:, 0], neg_points[:, 1], color=RED_COLOR, marker='*', s=marker_size, edgecolor='white',
            linewidth=LINEWIDTH)


def closetn(node, nodes):
    nodes = np.asarray(nodes)
    deltas = nodes - node
    dist_2 = np.einsum('ij,ij->i', deltas, deltas)
    return np.argmin(dist_2)

def create_workbook(name):
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'slice'
    ws['B1'] = '# green dots of best'
    ws['C1'] = '# red dots of best '
    ws['D1'] = 'SD of green of best '
    ws['E1'] = 'SD of red of best'
    ws['F1'] = 'best score'

    for i in range(9):
        coun = 1
        for col in ws.iter_cols(min_row=1, max_row=1, max_col=12 + i * 5, min_col=7 + i * 5):
            if coun == 1:
                ws[col[0].coordinate] = '# green dots of ' + str(i + 2)
            elif coun == 2:
                ws[col[0].coordinate] = '# red dots of ' + str(i + 2)
            elif coun == 3:
                ws[col[0].coordinate] = 'SD of X of ' + str(i + 2)
            elif coun == 4:
                ws[col[0].coordinate] = 'SD of Y of ' + str(i + 2)
            elif coun == 5:
                ws[col[0].coordinate] = 'score of ' + str(i + 2)
            coun += 1

    if not os.path.exists(name):
        os.makedirs(name)
        os.makedirs(os.path.join(name, "masks"))
        os.makedirs(os.path.join(name, "points"))
        os.makedirs(os.path.join(name, "sorts"))
        os.makedirs(os.path.join(name, "eachround"))
        os.makedirs(os.path.join(name, "scores"))
    
    return wb, ws