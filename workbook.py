from openpyxl import Workbook, load_workbook
import os
import time
import numpy as np

wb, ws, serv, tim, name, t = None, None, None, None, None, None

def open_workbook(continue_prev=False, in_name=None):
    global name, t
    
    if in_name:
        name = in_name
    else:
        name = input("what is your name?\n")
        
    if not continue_prev:
        while True:
            first = input("Do you want to load previous work? (y/n)\n").lower()
            if first in ('y', 'n'):
                break
            print("Please choose 'y' for yes or 'n' for no.")

        if first == 'n':
            c = create_workbook()
        else:
            c = open_existing_workbook()
    else:
        c = open_existing_workbook()

    t = time.time()
    return c

def update_survey():
    global serv
    while True:
        ans = input("Do you think the ground truth mask was suboptimal? (i.e. are SAM's results qualitatively better) y or n\n").lower()
        if ans in ('y', 'n'):
            break
        print("Please choose 'y' for yes or 'n' for no.")
    ans = 1 if ans=="y" else 0 
    serv=np.append(serv,ans)

def update_sample(c):
    # samples name on excel
    ws['A' + str(c + 2)] = str(c)

def save_workbook():
    wb.save(os.path.join(name, name + '.xlsx'))
    file = open(os.path.join(name, "time.txt"), 'w')
    file.write(str(float(tim) + (time.time() - t)))
    np.save(os.path.join(name,"servey.npy"),serv)
    file.close()

def save_metrics(c, score, ng, nr, stdx, stdy, gp, rp, msk, round):
    indx   = np.argsort(-np.array(score))
    sscore = np.array(score)[indx]
    sng    = np.array(ng)[indx]
    snr    = np.array(nr)[indx]
    sstdx  = np.array(stdx)[indx]
    sstdy  = np.array(stdy)[indx]

    for i in range(len(score)):
        coun = 1
        for col in ws.iter_cols(min_row=c + 2, max_row=c + 2, max_col=6 + i * 5, min_col=2 + i * 5):
            if coun == 1:
                ws[col[0].coordinate] = sng[i]
            elif coun == 2:
                ws[col[0].coordinate] = snr[i]
            elif coun == 3:
                ws[col[0].coordinate] = sstdx[i]
            elif coun == 4:
                ws[col[0].coordinate] = sstdy[i]
            elif coun == 5:
                ws[col[0].coordinate] = sscore[i]
            coun += 1

    np.save(os.path.join(name, "points"   , str(c) + "_green"), np.array(gp, dtype=object))
    np.save(os.path.join(name, "points"   , str(c) + "_red"  ), np.array(rp, dtype=object))
    np.save(os.path.join(name, "masks"    , str(c) + "_mask" ), np.array(msk))
    np.save(os.path.join(name, "sorts"    , str(c) + "_sort" ), indx)
    np.save(os.path.join(name, "scores"   , str(c) + "score" ), score)
    np.save(os.path.join(name, "eachround", str(c) + "_"     ), round)


def create_workbook():
    global wb, ws, tim, serv
    c        = 0
    tim      = 0
    serv     = np.array([])
    wb       = Workbook()
    ws       = wb.active

    ws['A1'] = 'slice'
    ws['B1'] = '# green dots of best'
    ws['C1'] = '# red dots of best '
    ws['D1'] = 'SD of green of best '
    ws['E1'] = 'SD of red of best'
    ws['F1'] = 'best score'

    for i in range(9):
        coun = 1
        for col in ws.iter_cols(min_row=1, max_row=1, max_col=12 + i * 5, min_col=7 + i * 5):
            if coun == 1:
                ws[col[0].coordinate] = '# green dots of ' + str(i + 2)
            elif coun == 2:
                ws[col[0].coordinate] = '# red dots of ' + str(i + 2)
            elif coun == 3:
                ws[col[0].coordinate] = 'SD of X of ' + str(i + 2)
            elif coun == 4:
                ws[col[0].coordinate] = 'SD of Y of ' + str(i + 2)
            elif coun == 5:
                ws[col[0].coordinate] = 'score of ' + str(i + 2)
            coun += 1

    if not os.path.exists(name):
        os.makedirs(name)
        os.makedirs(os.path.join(name, "masks"))
        os.makedirs(os.path.join(name, "points"))
        os.makedirs(os.path.join(name, "sorts"))
        os.makedirs(os.path.join(name, "eachround"))
        os.makedirs(os.path.join(name, "scores"))
    return c

def open_existing_workbook():
    global wb, ws, tim, serv
    wb = load_workbook(os.path.join(name, name + ".xlsx"))
    ws = wb.active
    c = len(os.listdir(os.path.join(name, "masks")))
    f = open(os.path.join(name, "time.txt"), 'r')
    serv=np.load(os.path.join(name,"servey.npy")) if os.path.exists(os.path.join(name,"servey.npy")) else np.array([])
    tim = f.readline()
    f.close()
    return c

